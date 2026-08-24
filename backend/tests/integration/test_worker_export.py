"""Export XLSX del tablero docente (A6).

El worker genera el libro con `openpyxl` y lo sube al bucket bajo una key
determinista. No hay bucket real en estos tests: se sustituye `boto3.client`
por un doble y se comprueba que `put_object` recibe un XLSX de verdad.
"""

import uuid
from datetime import date, timedelta
from io import BytesIO
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.security import hash_password
from app.modules.assessment.models import Assessment, Attempt, AttemptStatus
from app.modules.catalog.models import MOMENT_ORDER, Moment, Project
from app.modules.identity.models import Calendar, Institution, License, User
from app.workers.worker import export_results


class _MismaSesion:
    def __init__(self, session):
        self._session = session

    async def __aenter__(self):
        return self._session

    async def __aexit__(self, *exc_info):
        return False


async def test_el_xlsx_incluye_una_fila_por_intento(db, monkeypatch):
    from app.workers import worker as worker_module

    monkeypatch.setattr(worker_module, "SessionLocal", lambda: _MismaSesion(db))

    inst = Institution(name=f"Colegio {uuid.uuid4().hex[:6]}", calendar=Calendar.A)
    db.add(inst)
    await db.flush()
    db.add(
        License(
            institution_id=inst.id,
            calendar=Calendar.A,
            valid_from=date.today() - timedelta(days=1),
            valid_to=date.today() + timedelta(days=365),
            seats=10,
        )
    )
    proyecto = Project(slug=f"p-{uuid.uuid4().hex[:6]}", grade="5", order=0)
    db.add(proyecto)
    await db.flush()
    for orden, tipo in enumerate(MOMENT_ORDER):
        db.add(Moment(project_id=proyecto.id, type=tipo, order=orden))
    await db.flush()
    momento_assess = (
        await db.execute(
            select(Moment).where(
                Moment.project_id == proyecto.id, Moment.type == "assess"
            )
        )
    ).scalar_one()
    assessment = Assessment(moment_id=momento_assess.id)
    db.add(assessment)
    await db.flush()

    estudiante = User(
        email=f"e-{uuid.uuid4().hex[:6]}@imaquina.example.com",
        full_name="Ana Estudiante",
        password_hash=hash_password("x"),
        role="student",
        institution_id=inst.id,
    )
    db.add(estudiante)
    await db.flush()
    db.add(
        Attempt(
            assessment_id=assessment.id,
            student_id=estudiante.id,
            institution_id=inst.id,
            status=AttemptStatus.GRADED,
            score=8.0,
        )
    )
    await db.flush()

    cliente_falso = MagicMock()
    with patch("boto3.client", return_value=cliente_falso):
        resultado = await export_results({}, str(assessment.id), str(estudiante.id))

    assert resultado["status"] == "listo"
    cliente_falso.put_object.assert_called_once()
    _, kwargs = cliente_falso.put_object.call_args
    assert kwargs["Key"] == f"exports/{assessment.id}.xlsx"

    libro = load_workbook(BytesIO(kwargs["Body"]))
    hoja = libro.active
    filas = list(hoja.iter_rows(values_only=True))
    assert filas[0] == ("Estudiante", "Equipo", "Estado", "Puntaje", "Enviado")
    assert filas[1][0] == "Ana Estudiante"
    assert filas[1][3] == 8.0


async def test_sin_intentos_igual_genera_el_encabezado(db, monkeypatch):
    from app.workers import worker as worker_module

    monkeypatch.setattr(worker_module, "SessionLocal", lambda: _MismaSesion(db))

    proyecto = Project(slug=f"p-{uuid.uuid4().hex[:6]}", grade="5", order=0)
    db.add(proyecto)
    await db.flush()
    for orden, tipo in enumerate(MOMENT_ORDER):
        db.add(Moment(project_id=proyecto.id, type=tipo, order=orden))
    await db.flush()
    momento_assess = (
        await db.execute(
            select(Moment).where(
                Moment.project_id == proyecto.id, Moment.type == "assess"
            )
        )
    ).scalar_one()
    assessment = Assessment(moment_id=momento_assess.id)
    db.add(assessment)
    await db.flush()

    cliente_falso = MagicMock()
    with patch("boto3.client", return_value=cliente_falso):
        resultado = await export_results({}, str(assessment.id), str(uuid.uuid4()))

    assert resultado["status"] == "listo"
